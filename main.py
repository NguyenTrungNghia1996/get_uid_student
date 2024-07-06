import time
import openpyxl
import requests

from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from halo import Halo

from progress.bar import IncrementalBar

from concurrent.futures import ThreadPoolExecutor, as_completed

root = Tk()
root.attributes("-topmost", True)
root.withdraw()

user = "961005057"
password = "Speakwell@2023!"
delay = 3

apiUrl = "https://dynamic.classin.com/saasajax/student.ajax.php?action=getSchoolStudentListByPage"


class Student:
    def __init__(self, _id, studId, studentUid, mobile, email, studentName):
        self._id = _id
        self.studId = studId
        self.studentUid = studentUid
        self.mobile = mobile
        self.email = email
        self.studentName = studentName

    def __str__(self):
        return f"{self._id}, {self.studId}, {self.studentUid},{self.mobile},{self.email},{self.studentName}"


def add_student(students, _id, studId, studentUid, mobile, email, studentName):
    student = Student(_id, studId, studentUid, mobile, email, studentName)
    students.append(student)


def edit_student(students, _id, new_studId, new_studentUid, new_mobile, new_email, new_studentName):
    for student in students:
        if student.id == _id:
            student.studId = new_studId
            student.studentUid = new_studentUid
            student.mobile = new_mobile
            student.email = new_email
            student.studentName = new_studentName
            break


def remove_decimal_zero(num_str):
    if num_str.endswith('.0'):
        return num_str[:-2]
    return num_str


def select_file():
    # root = tk.Tk()
    # root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path


def read_first_column(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    students = []
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        students.append(remove_decimal_zero(str(row[0])))
    return students


def read_excel_file():
    spinner = Halo(text='Đang chọn file excel nguồn', spinner='dots')
    spinner.start()
    file_path = select_file()
    if file_path:
        data_exel = read_first_column(file_path)
        number_id = len(data_exel)
        spinner.stop()
        spinner.succeed("Đã nhập " + str(number_id) + " ID Học sinh")
        return data_exel
    else:
        spinner.stop()
        spinner.fail("Nhập File excel không thành công")


def save_to_excel(students):
    spinner = Halo(text='Chọn chỗ lưu file Excel', spinner='dots')
    spinner.start()
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if not file_path:
        print("\n Đã hủy lưu file.")
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"

    headers = ["ID", "StudID", "StudentUID", "Mobile", "Email", "StudentName"]
    sheet.append(headers)

    for student in students:
        sheet.append(
            [student._id, student.studId, student.studentUid, student.mobile, student.email, student.studentName])

    workbook.save(file_path)
    spinner.stop()
    spinner.succeed(f"\nFile Excel đã được lưu tại: {file_path}")


def signin():
    spinner = Halo(text='Đang đăng nhập...', spinner='dots')
    spinner.start()

    try:
        driver = webdriver.Chrome()
        driver.get("https://www.classin.com/login/")
        user_input = driver.find_element(By.ID, "phone")
        password_input = driver.find_element(By.ID, "password")
        user_input.send_keys(user)
        password_input.send_keys(password)
        select_area = driver.find_element(By.ID, 'area')
        select = Select(select_area)
        select.select_by_value('84')
        time.sleep(3)
        summit_button = driver.find_element(By.ID, 'mySubmit')
        summit_button.click()
        time.sleep(10)
        messagebox.showinfo(title='Captcha verification', message='Hoàn thành đăng nhập sau đó mới click OK',
                            parent=root)
        all_cookies = driver.get_cookies()
        cookies_dict = {}
        for cookie in all_cookies:
            cookies_dict[cookie['name']] = cookie['value']
        spinner.stop()
        spinner.succeed("Đăng nhập thành công !")
        return cookies_dict
    except Exception as e:
        spinner.stop()
        spinner.fail("Lỗi đăng nhập !", e)


def get_uid_student(_id, cookies):
    params = {
        "page": 1,
        "perpage": 20,
        "searchKey": _id,
        "isdel": 0
    }
    try:
        response = requests.post(apiUrl, data=params, cookies=cookies)
        response.raise_for_status()
        # return response
        res = response.json()
        if res["error_info"]["errno"] == 1 and len(res["data"]["studentList"]) > 0:
            data = res["data"]["studentList"][0]
            student = Student(_id, data["studId"] or "", data["studentUid"] or "", data["mobile"] or "",
                              data["email"] or "", data["studentName"] or "")
            return student
        else:
            student = Student(_id, "", "", "", "", "")
            return student
    except requests.RequestException as e:
        print(f"Error fetching {_id}: {e}")
        return None


def fetch_all(data_id, cookies):
    results = []
    with ThreadPoolExecutor(max_workers=50) as executor:
        futures = [executor.submit(get_uid_student, item, cookies) for item in data_id]
        with IncrementalBar('Processing...', max=len(data_id), suffix='%(index)d/%(max)d - %(elapsed_td)s') as bar:
            for future in as_completed(futures):
                result = future.result()
                if result is not None:
                    results.append(result)
                bar.next()

    return results


def main():
    cookies = signin()
    data_id = read_excel_file()
    students = fetch_all(data_id, cookies)
    save_to_excel(students)


if __name__ == "__main__":
    main()
